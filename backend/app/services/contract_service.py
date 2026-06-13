"""合同服务层：预览、生成、导出、下载、批量导入"""

import os
import uuid
import zipfile
from datetime import datetime

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.contract import Contract
from app.models.template import Template, TemplateVersion
from app.utils.doc_generator import generate_docx, preview_docx


async def preview_contract(
    db: AsyncSession,
    template_id: uuid.UUID,
    variables: dict[str, str],
    template_version_id: uuid.UUID | None = None,
) -> str:
    """预览：替换变量后返回文本"""
    template_version = await _get_template_version(
        db, template_id, template_version_id
    )
    if not template_version:
        raise ValueError("模板版本不存在")

    return preview_docx(template_version.file_path, variables)


async def generate_contract(
    db: AsyncSession,
    title: str,
    template_id: uuid.UUID,
    variables: dict[str, str],
    project_id: uuid.UUID | None = None,
    template_version_id: uuid.UUID | None = None,
    user_id: uuid.UUID | None = None,
) -> Contract:
    """生成合同文档"""
    template_version = await _get_template_version(
        db, template_id, template_version_id
    )
    if not template_version:
        raise ValueError("模板版本不存在")

    # 生成 DOCX 文件
    output_dir = os.path.join("./uploads", "contracts")
    os.makedirs(output_dir, exist_ok=True)

    output_filename = f"{uuid.uuid4().hex}.docx"
    output_path = os.path.join(output_dir, output_filename)

    generate_docx(template_version.file_path, variables, output_path)

    # 创建 Contract 记录（自动归档）
    now = datetime.utcnow().isoformat()
    contract = Contract(
        title=title,
        project_id=project_id,
        template_id=template_id,
        template_version_id=template_version.id,
        variables=variables,
        file_path=output_path,
        status="archived",
        archived_at=datetime.utcnow(),
        status_history=[
            {"status": "draft", "at": now},
            {"status": "archived", "at": now},
        ],
        created_by=user_id,
    )
    db.add(contract)
    await db.flush()

    return contract


async def get_contract(db: AsyncSession, contract_id: uuid.UUID) -> Contract | None:
    """获取合同详情"""
    result = await db.execute(
        select(Contract).where(Contract.id == contract_id)
    )
    return result.scalar_one_or_none()


async def list_contracts(
    db: AsyncSession,
    page: int = 1,
    page_size: int = 20,
    project_id: uuid.UUID | None = None,
    status: str | None = None,
) -> tuple[list[Contract], int]:
    """合同列表"""
    query = select(Contract)

    if project_id:
        query = query.where(Contract.project_id == project_id)
    if status:
        query = query.where(Contract.status == status)

    count_query = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_query)).scalar() or 0

    query = query.order_by(Contract.created_at.desc())
    query = query.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)

    return list(result.scalars().all()), total


async def delete_contract(db: AsyncSession, contract_id: uuid.UUID) -> bool:
    """删除合同"""
    contract = await get_contract(db, contract_id)
    if not contract:
        return False

    # 删除文件
    if contract.file_path and os.path.exists(contract.file_path):
        os.remove(contract.file_path)

    await db.delete(contract)
    await db.flush()
    return True


async def export_contract(
    db: AsyncSession,
    contract_id: uuid.UUID,
    format: str = "word",
) -> str | None:
    """导出合同文件，返回文件路径"""
    contract = await get_contract(db, contract_id)
    if not contract:
        return None

    if format == "word" or format == "docx":
        return contract.file_path
    elif format == "pdf":
        if contract.file_path_pdf and os.path.exists(contract.file_path_pdf):
            return contract.file_path_pdf

        # 即时转换 DOCX → PDF
        from app.utils.pdf_converter import convert_docx_to_pdf, is_libreoffice_available
        if not is_libreoffice_available():
            return None

        if contract.file_path and os.path.exists(contract.file_path):
            pdf_dir = os.path.dirname(contract.file_path)
            pdf_path = convert_docx_to_pdf(contract.file_path, pdf_dir)
            if pdf_path:
                contract.file_path_pdf = pdf_path
                await db.flush()
                return pdf_path

        return None

    return contract.file_path


async def parse_excel(
    excel_path: str,
) -> tuple[list[str], list[dict[str, str]]]:
    """解析 Excel 文件，返回表头和每行数据（不生成合同）"""
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 1:
        wb.close()
        raise ValueError("Excel 文件为空")

    headers = [str(h).strip() if h else "" for h in rows[0]]
    data_rows: list[dict[str, str]] = []

    for row in rows[1:]:
        row_data: dict[str, str] = {}
        for col_idx, value in enumerate(row):
            if col_idx >= len(headers) or not headers[col_idx]:
                continue
            row_data[headers[col_idx]] = str(value).strip() if value is not None else ""
        if row_data:
            data_rows.append(row_data)

    wb.close()
    return headers, data_rows


async def batch_generate_from_rows(
    db: AsyncSession,
    project_id: uuid.UUID,
    rows: list[dict[str, str]],
    selected_indices: list[int],
    user_id: uuid.UUID | None = None,
) -> list[Contract]:
    """从选中的行数据批量生成合同（每个模板各生成一份）"""
    from app.models.project import Project

    # 获取项目及其模板
    result = await db.execute(
        select(Project).where(Project.id == project_id)
    )
    project = result.scalar_one_or_none()
    if not project:
        raise ValueError("项目不存在")

    # 获取项目关联的模板
    from sqlalchemy.orm import selectinload
    result = await db.execute(
        select(Project).options(selectinload(Project.templates)).where(Project.id == project_id)
    )
    project = result.scalar_one_or_none()
    if not project:
        raise ValueError("项目不存在")

    contracts: list[Contract] = []
    output_dir = os.path.join("./uploads", "contracts")
    os.makedirs(output_dir, exist_ok=True)

    for idx in selected_indices:
        if idx < 0 or idx >= len(rows):
            continue
        row_vars = rows[idx]

        for tmpl in project.templates:
            template_version = await _get_template_version(db, tmpl.id)
            if not template_version:
                continue

            # 合并：Step 2 手动填的变量 + Excel 行的变量（Excel 优先覆盖）
            variables = row_vars

            # 生成有意义的文件名
            var_summary = "_".join(v for v in list(variables.values())[:3] if v)
            safe_summary = "".join(c for c in var_summary if c.isalnum() or c in "_-")[:50]
            output_filename = f"{tmpl.name}_{safe_summary}_{uuid.uuid4().hex[:8]}.docx"
            output_path = os.path.join(output_dir, output_filename)

            generate_docx(template_version.file_path, variables, output_path)

            # 构建标题
            row_label = f"第{idx + 1}行"
            title = f"{project.name} - {tmpl.name} - {row_label}"
            if safe_summary:
                title += f" - {safe_summary}"

            now = datetime.utcnow().isoformat()
            contract = Contract(
                title=title,
                project_id=project_id,
                template_id=tmpl.id,
                template_version_id=template_version.id,
                variables=variables,
                file_path=output_path,
                status="archived",
                archived_at=datetime.utcnow(),
                status_history=[
                    {"status": "draft", "at": now},
                    {"status": "archived", "at": now},
                ],
                created_by=user_id,
            )
            db.add(contract)
            contracts.append(contract)

    await db.flush()
    return contracts


async def _get_template_version(
    db: AsyncSession,
    template_id: uuid.UUID,
    template_version_id: uuid.UUID | None = None,
) -> TemplateVersion | None:
    """获取模板版本，默认使用 master 版本"""
    if template_version_id:
        result = await db.execute(
            select(TemplateVersion).where(TemplateVersion.id == template_version_id)
        )
        return result.scalar_one_or_none()

    # 获取 master 版本
    result = await db.execute(
        select(TemplateVersion).where(
            TemplateVersion.template_id == template_id,
            TemplateVersion.is_master == True,
        )
    )
    return result.scalar_one_or_none()


async def batch_generate_from_excel(
    db: AsyncSession,
    template_id: uuid.UUID,
    excel_path: str,
    project_id: uuid.UUID | None = None,
    title_column: str = "标题",
    user_id: uuid.UUID | None = None,
) -> list[Contract]:
    """从 Excel 文件批量生成合同

    Excel 格式：第一行为变量名（表头），后续每行为一组变量值。
    可选包含"标题"列指定合同标题，否则用行号生成。
    """
    template_version = await _get_template_version(db, template_id)
    if not template_version:
        raise ValueError("模板版本不存在")

    # 读取 Excel
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        wb.close()
        raise ValueError("Excel 至少需要表头行和一行数据")

    # 表头即变量名
    headers = [str(h).strip() if h else "" for h in rows[0]]
    contracts: list[Contract] = []

    output_dir = os.path.join("./uploads", "contracts")
    os.makedirs(output_dir, exist_ok=True)

    for row_idx, row in enumerate(rows[1:], start=2):
        # 构建变量映射
        variables: dict[str, str] = {}
        row_title = None
        for col_idx, value in enumerate(row):
            if col_idx >= len(headers) or not headers[col_idx]:
                continue
            header = headers[col_idx]
            cell_value = str(value).strip() if value is not None else ""

            if header == title_column:
                row_title = cell_value
            else:
                variables[header] = cell_value

        if not variables:
            continue

        # 生成标题
        title = row_title or f"批量生成_{row_idx}"

        # 生成 DOCX
        output_filename = f"{uuid.uuid4().hex}.docx"
        output_path = os.path.join(output_dir, output_filename)
        generate_docx(template_version.file_path, variables, output_path)

        now = datetime.utcnow().isoformat()
        contract = Contract(
            title=title,
            project_id=project_id,
            template_id=template_id,
            template_version_id=template_version.id,
            variables=variables,
            file_path=output_path,
            status="archived",
            archived_at=datetime.utcnow(),
            status_history=[
                {"status": "draft", "at": now},
                {"status": "archived", "at": now},
            ],
            created_by=user_id,
        )
        db.add(contract)
        contracts.append(contract)

    await db.flush()
    wb.close()
    return contracts


def build_zip(contracts: list[Contract], output_dir: str = "./uploads/zip") -> str:
    """将多个合同文件打包为 zip，返回 zip 文件路径"""
    os.makedirs(output_dir, exist_ok=True)
    zip_filename = f"contracts_{uuid.uuid4().hex[:8]}.zip"
    zip_path = os.path.join(output_dir, zip_filename)

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for contract in contracts:
            if contract.file_path and os.path.exists(contract.file_path):
                # 用有意义的文件名
                arcname = os.path.basename(contract.file_path)
                zf.write(contract.file_path, arcname)

    return zip_path