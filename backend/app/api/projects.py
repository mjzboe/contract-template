"""项目管理 API 路由"""

import os
import uuid

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.dependencies import require_role
from app.models.user import User
from app.schemas.project import (
    DeduplicatedVariablesResponse,
    ProjectCreate,
    ProjectListResponse,
    ProjectResponse,
    ProjectUpdate,
    VariableInfoResponse,
)
from app.schemas.template import VariableInfoResponse as TemplateVarResp
from app.services import project_service

router = APIRouter(prefix="/projects", tags=["项目管理"])


@router.post("", response_model=ProjectResponse)
async def create_project(
    data: ProjectCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role("user", "template_admin", "approver", "super_admin")),
):
    """创建项目，关联模板并自动变量去重"""
    project = await project_service.create_project(db, data, current_user.id)
    await db.commit()
    return ProjectResponse.model_validate(project)


@router.get("", response_model=ProjectListResponse)
async def list_projects(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    keyword: str | None = None,
    status: str | None = None,
    db: AsyncSession = Depends(get_db),
):
    """项目列表"""
    projects, total = await project_service.list_projects(
        db, page, page_size, keyword, status
    )
    return ProjectListResponse(
        items=[ProjectResponse.model_validate(p) for p in projects],
        total=total,
        page=page,
        page_size=page_size,
    )


@router.get("/{project_id}", response_model=ProjectResponse)
async def get_project(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
):
    """项目详情"""
    project = await project_service.get_project(db, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    return ProjectResponse.model_validate(project)


@router.put("/{project_id}", response_model=ProjectResponse)
async def update_project(
    project_id: uuid.UUID,
    data: ProjectUpdate,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_role("user", "template_admin", "approver", "super_admin")),
):
    """更新项目信息（如更新模板列表会自动重新去重）"""
    project = await project_service.update_project(db, project_id, data)
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    await db.commit()
    return ProjectResponse.model_validate(project)


@router.delete("/{project_id}")
async def delete_project(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_role("super_admin")),
):
    """删除项目"""
    success = await project_service.delete_project(db, project_id)
    if not success:
        raise HTTPException(status_code=404, detail="项目不存在")
    await db.commit()
    return {"message": "删除成功"}


@router.get("/{project_id}/deduplicated-variables", response_model=DeduplicatedVariablesResponse)
async def get_deduplicated_variables(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
):
    """获取项目的跨模板去重变量，含来源映射"""
    result = await project_service.get_deduplicated_variables(db, project_id)
    if not result:
        raise HTTPException(status_code=404, detail="项目不存在")

    return DeduplicatedVariablesResponse(
        project_id=result["project_id"],
        template_count=result["template_count"],
        total_variables_before_dedup=result["total_variables_before_dedup"],
        total_variables_after_dedup=result["total_variables_after_dedup"],
        variables=[TemplateVarResp(**v) for v in result["variables"]],
        variable_sources=result["variable_sources"],
    )


@router.get("/{project_id}/excel-template")
async def download_excel_template(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_role("user", "template_admin", "approver", "super_admin")),
):
    """根据项目去重变量动态生成 Excel 导入模板"""
    result = await project_service.get_deduplicated_variables(db, project_id)
    if not result:
        raise HTTPException(status_code=404, detail="项目不存在")

    variables = result["variables"]
    if not variables:
        raise HTTPException(status_code=400, detail="项目无变量，无法生成模板")

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "变量导入"

    # 第一行：变量名（作为表头，也是 Excel 导入时的列名）
    for col, var in enumerate(variables, 1):
        display_name = var.get("display_name") or var.get("name", "")
        var_name = var.get("name", "")
        ws.cell(row=1, column=col, value=var_name)

        # 设置列宽
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max(len(display_name) * 2 + 4, 16)

    # 第二行：显示名提示行（灰色斜体，帮助用户理解每列含义）
    for col, var in enumerate(variables, 1):
        display_name = var.get("display_name") or var.get("name", "")
        cell = ws.cell(row=2, column=col, value=display_name)
        cell.font = cell.font.copy(italic=True, color="999999")

    # 第三行：默认值（如有）
    for col, var in enumerate(variables, 1):
        default_val = var.get("default_value", "")
        if default_val:
            ws.cell(row=3, column=col, value=default_val)

    # 保存到临时文件
    from app.config import settings
    template_dir = os.path.join(settings.UPLOAD_DIR, "excel_templates")
    os.makedirs(template_dir, exist_ok=True)
    file_path = os.path.join(template_dir, f"template_{project_id}.xlsx")
    wb.save(file_path)

    project_name = result.get("project_id", str(project_id))
    filename = f"导入模板_{project_name}.xlsx"

    return FileResponse(
        path=file_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
